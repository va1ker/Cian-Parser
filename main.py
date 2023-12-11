from openpyxl import Workbook, load_workbook
import os
from settings import FILE_NAME, COLUMN_NAMES
import requests
from bs4 import BeautifulSoup
from alive_progress import alive_bar


def flat_parser(link):
    req = requests.get(link).content
    flat_data = BeautifulSoup(req, "html.parser")
    flat_name = (
        flat_data.find("div", class_="a10a3f92e9--container--pWxZo").find("h1").text
        or "-"
    )
    flat_price = (
        flat_data.find("div", class_="a10a3f92e9--amount--ON6i1")
        .find("span")
        .text.replace("&nbsp;", "")
        .replace("\xa0", "")
        or "-"
    )
    flat_type = flat_data.find("span", text="Тип жилья").findNext("span").text or "-"
    flat_total = (
        flat_data.find("span", text="Общая площадь")
        .findNext("span")
        .text.replace("\xa0", "")
        or "-"
    )
    flat_addres = flat_data.find(
        "div", class_="a10a3f92e9--address-line--GRDTb"
    ).find_all("a")
    flat_addres = ",".join([addres.text for addres in flat_addres])
    data = {
        "Link": link.strip("\n"),
        "Name": flat_name,
        "Price": flat_price,
        "Addres": flat_addres,
        "Total": flat_total,
        "Type": flat_type,
        "Watched": False,
        "Accepted": 0,
    }
    return data


def link_exists_in_excel(filename, link_to_check):
    workbook = load_workbook(filename)
    sheet = workbook.active

    existing_links = [
        str(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1)
    ]

    if link_to_check in existing_links:
        return True
    else:
        return False


def write_data_to_excel(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    row = []
    for item in data:
        row.append(data[item])
    sheet.append(row)

    workbook.save(filename)


def check_table_exist():
    wb = Workbook()
    ws = wb.active
    ws.title = "Flats"
    if not os.path.exists(FILE_NAME):
        print("Таблица не найдена, создаем...")
        ws.append(COLUMN_NAMES)
        wb.save(filename=FILE_NAME)
        print(f"Файл {FILE_NAME} создан.")
    else:
        ws.append(COLUMN_NAMES)
        wb.save(filename=FILE_NAME)
        print(f"Таблица {FILE_NAME} найдена.")
    return True


if __name__ == "__main__":
    check_table_exist()
    with open("links.txt") as file:
        links = file.readlines()
        with alive_bar(len(links)) as bar:
            for link in links:
                flag = link_exists_in_excel(FILE_NAME, link)
                if not flag:
                    data = flat_parser(link)
                    write_data_to_excel(FILE_NAME, data)
                bar()
